import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\u724504\\OneDrive - Finastra\Desktop\\Personal\\Udemy\\Python-selenium\\PYExcel_demo\\demo_sheet.xlsx")

#go to the active sheet in workbook
ActiveSheet = workbook.active


#print data of any cell from the sheet
ProcessOne = ActiveSheet.cell(row=2, column=3).value
print(ProcessOne)
ProcessTwo = ActiveSheet["C1"]
print(ProcessTwo.value)


#write data to some cell and get it's value
ActiveSheet.cell(row=8, column=10).value = "test from code"
print(ActiveSheet.cell(row=8, column=10).value)


#get the max count of rows and coloumns
MaxRows = ActiveSheet.max_row # As long as any row is referenced in the code prior to this line, it will be considered in this count
MaxColoumns = ActiveSheet.max_column # Same as above, applied on coloumns.
print(MaxRows, MaxColoumns)


#iterate over all the values in the sheet for Testcase3, leaving the value "Testcase3)
for i in range(1, MaxRows+1):
    if ActiveSheet.cell(row=i, column=1).value == "Testcase3":
        for j in range(2, MaxColoumns+1):
            if not ActiveSheet.cell(row=i, column=j).value == None:
                print(ActiveSheet.cell(row=i, column=j).value) #Note: if any cell has formula, the whole formula comes as it's value

#Save the same information as above in a dictionary
dic = {}
for i in range(1, MaxRows+1):
    if ActiveSheet.cell(row=i, column=1).value == "Testcase3":
        for j in range(2, MaxColoumns+1):
            if not ActiveSheet.cell(row=i, column=j).value == None:
                dic[ActiveSheet.cell(row=1, column=j).value] = ActiveSheet.cell(row=i, column=j).value


print(dic)
